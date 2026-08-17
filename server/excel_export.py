# -*- coding: utf-8 -*-
"""
Excel 报表导出：出车 / 加油 / 保养
基于 openpyxl 生成 .xlsx，返回 (bytes, 文件名)
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import db

HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center")


def _write_rows(ws, headers, rows):
    """写入表头 + 数据行，返回表头样式与列数。"""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    for row in rows:
        ws.append(row)


def _auto_width(ws, ncols, min_w=10, max_w=22):
    """根据表头长度与首行数据估算列宽。"""
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        # 表头宽度为基准
        w = len(str(ws.cell(1, col).value or ""))
        for r in range(2, min(ws.max_row, 8) + 1):
            v = ws.cell(r, col).value
            if v is not None:
                w = max(w, len(str(v)))
        ws.column_dimensions[letter].width = max(min_w, min(w * 1.6 + 4, max_w))


def _fmt(v, nd=2):
    """数字保留 nd 位小数，去掉多余 0。"""
    if v is None:
        return ""
    try:
        f = float(v)
    except (TypeError, ValueError):
        return v
    if f == int(f):
        return int(f)
    return round(f, nd)


def export_trips(date_from="", date_to=""):
    rows = db.report_trips_detail(date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "出车报表"
    headers = ["序号", "日期", "姓名", "车牌号", "出发地", "目的地", "提交时间"]
    data = [[i + 1, r["trip_date"], r["name"], r["plate"],
             r["origin"], r["destination"], r["created_at"]]
            for i, r in enumerate(rows)]
    _write_rows(ws, headers, data)
    if data:
        ws.append(["合计", "", "", "", "", "", "共 %d 次出车" % len(rows)])
        for cell in ws[ws.max_row]:
            cell.font = TOTAL_FONT
    _auto_width(ws, len(headers))
    fname = "金成峰司机-出车报表-%s.xlsx" % datetime.now().strftime("%Y%m%d")
    return _to_bytes(wb), fname


def export_refuels(date_from="", date_to=""):
    rows = db.report_refuels_detail(date_from, date_to)
    summary = db.report_refuels_summary(date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "加油报表"
    headers = ["序号", "日期", "姓名", "车牌号", "公里数", "行驶公里(km)",
               "油价(元/升)", "加油量(升)", "金额(元)", "油耗(升/百公里)", "提交时间"]
    data = [[i + 1, r["refuel_date"], r["name"], r["plate"], _fmt(r["odometer"]),
             _fmt(r["travel_km"]), _fmt(r["oil_price"], 2), _fmt(r["liters"], 2),
             _fmt(r["amount"], 2), _fmt(r["fuel_consumption"], 2), r["created_at"]]
            for i, r in enumerate(rows)]
    _write_rows(ws, headers, data)
    if data:
        ws.append(["合计", "", "", "", "", "", "",
                   "共 %d 次加油" % int(summary["cnt"]),
                   "总金额 %s 元" % _fmt(summary["total_amount"]),
                   "总升数 %s" % _fmt(summary["total_liters"]), ""])
        for cell in ws[ws.max_row]:
            cell.font = TOTAL_FONT
    _auto_width(ws, len(headers))
    fname = "金成峰司机-加油报表-%s.xlsx" % datetime.now().strftime("%Y%m%d")
    return _to_bytes(wb), fname


def export_maintenances(date_from="", date_to=""):
    rows = db.report_maintenances_detail(date_from, date_to)
    summary = db.report_maintenances_summary(date_from, date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "保养报表"
    headers = ["序号", "保养时间", "姓名", "车牌号", "保养项目", "费用(元)", "备注", "提交时间"]
    data = [[i + 1, r["maintain_time"], r["name"], r["plate"], r["items"],
             _fmt(r["cost"], 2), r["remark"], r["created_at"]]
            for i, r in enumerate(rows)]
    _write_rows(ws, headers, data)
    if data:
        ws.append(["合计", "", "", "", "", "总费用 %s 元" % _fmt(summary["total_cost"]),
                   "共 %d 次保养" % int(summary["cnt"]), ""])
        for cell in ws[ws.max_row]:
            cell.font = TOTAL_FONT
    _auto_width(ws, len(headers))
    fname = "金成峰司机-保养报表-%s.xlsx" % datetime.now().strftime("%Y%m%d")
    return _to_bytes(wb), fname


def _to_bytes(wb):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


EXPORTERS = {
    "trips": export_trips,
    "refuels": export_refuels,
    "maintenances": export_maintenances,
}


if __name__ == "__main__":
    db.init_db()
    for kind in EXPORTERS:
        data, fname = EXPORTERS[kind]()
        print(fname, len(data), "bytes")
